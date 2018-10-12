import os
import bcrypt
import numpy as np

from random import random
from datetime import datetime
from openpyxl import load_workbook
from bson.objectid import ObjectId
from bson.json_util import loads, dumps

# Aux functions for bottle server
from functions import *
from models import dengue

from bottle.ext.mongo import MongoPlugin
from bottle import Bottle, route, get, post, template, redirect, static_file, error, run, request, response, default_app

data_download_template = 'static/files/templates/data_download.xlsx'

APP_TITLE = 'Epidemiología | '

debugging = True

app = Bottle()
plugin = MongoPlugin(uri=os.environ['MONGODB_URI'], db="heroku_r18zcfb4", json_mongo=True, keyword='mongodb')

def is_auth(mongodb, cookie):
    return mongodb.sessions.count({'cookie':cookie}) > 0

def check_login(mongodb, user, password):
    userData = mongodb.users.find_one({'user':user})

    if(userData):
        if(bcrypt.checkpw(password.encode('utf8'), userData["hash"])):
            if(userData["role"] == "admin"):
                return 3
            elif(userData["role"] == "editor"):
                return 2
            else:
                return 1
    return 0

def tpl(name, page_title, error_msg=None, data=None):
    return template(name, title=APP_TITLE+page_title, error_msg=error_msg, data=data)

@app.route('/')
def main(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        if cookie[1] == "3":
            return tpl('admin/dashboard', 'Admin | Dashboard')
        elif cookie[1] == "1" or cookie[1] == "2":
            return tpl('dashboard', 'Dashboard', data=getDashboard(mongodb, cookie))

    return tpl('home', 'Inicio')

@app.route('/ingreso/')
def show_login(mongodb):
    if is_auth(mongodb, request.get_cookie('auth')):
        redirect('/')
    else:
        try:
            error_flag = request.query['login_error']
            error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>¡Usuario o contraseña incorrecta!</li></ul></div>'
        except KeyError:
            error_flag = False
            error_msg = ''

        return tpl('login', 'Ingreso', error_msg)

@app.route('/simulador/')
def show_simulator(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie}, {'user':1})['user']
        return tpl('simulator', 'Simulador', data=getParameters(mongodb, user))
    else:
        redirect('/')

@app.route('/datos/')
def show_data(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie}, {'user':1})['user']

        try:
            error_flag = request.query['data_error']
            if int(error_flag) > 0:
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>El archivo que cargaste no es compatible.</li></ul></div>'
            elif int(error_flag) == -1:
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>No seleccionaste ningún archivo.</li></ul></div>'
            else:
                error_msg = '<div style="display:block;" class="ui success message"><ul class="list"><li>El archivo fue cargado correctamente.</li></ul></div>'

        except KeyError:
            error_flag = False
            error_msg = ''

        return tpl('data', 'Módulo de datos', error_msg, {'cookie':cookie, 'cases':getDataByWeeks(mongodb, user)})
    else:
        redirect('/')

@app.route('/canal/')
def show_channel(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie},{'user':1})['user']
        data = {"user":user, "channels":getChannelYears(mongodb, user)}
        return tpl('channel', 'Canal endémico', data=data)
    else:
        redirect('/')

@app.route('/mapa/')
def show_map(mongodb):
    if is_auth(mongodb, request.get_cookie('auth')):
        return tpl('map', 'Mapa')
    else:
        redirect('/')

@app.route('/prediccion/')
def show_map(mongodb):
    if is_auth(mongodb, request.get_cookie('auth')):
        return tpl('prediction', 'Predicción')
    else:
        redirect('/')

@app.route('/riesgo/')
def show_risk_map(mongodb):
    if is_auth(mongodb, request.get_cookie('auth')):
        return tpl('risk', 'Mapa de riesgo')
    else:
        redirect('/')

@app.route('/riesgo/comunas/layers.json')
def map_risk(mongodb):
    cookie = request.get_cookie('auth')
    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie},{'user':1})['user']

        with open('static/files/geojson/%s_risk.json'%user, 'w') as tempfile:
            data = getRiskColors(mongodb, user)
            tempfile.write(dumps(data))
            response = static_file('%s_risk.json'%user, root='static/files/geojson')
            response.set_header("Cache-Control", "public, max-age=0")              ## TO CHANGE
            response.set_header("Content-Type", "application/json")
            return response
    else:
        redirect('/')

@app.route('/mapa/comunas/layers.json')
def map_layers(mongodb):
    cookie = request.get_cookie('auth')
    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie},{'user':1})['user']

        with open('static/files/geojson/%s_layers.json'%user, 'w') as tempfile:
            data = mongodb[user].layers.find_one()
            data.pop('_id', None)
            tempfile.write(dumps(data))
            response = static_file('%s_layers.json'%user, root='static/files/geojson')
            response.set_header("Cache-Control", "public, max-age=0")              ## TO CHANGE
            response.set_header("Content-Type", "application/json")
            return response
    else:
        redirect('/')

@app.route('/mapa/casos/cases.json')
def map_cases(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie},{'user':1})['user']

        with open('static/files/geojson/%s_cases.json'%user, 'w') as tempfile:
            tempfile.write(dumps(mongodb[user].byWeeks.geo.find_one()))
            response = static_file('%s_cases.json'%user, root='static/files/geojson')
            response.set_header("Cache-Control", "public, max-age=0")              ## TO CHANGE
            response.set_header("Content-Type", "application/json")
            return response
    else:
        redirect('/')

@app.get('/datos/cargar')
def redirect_to_data(mongodb):
    if is_auth(mongodb, request.get_cookie('auth')):
        redirect('/datos/')
    else:
        redirect('/')

@app.post('/datos/cargar')
def do_upload(mongodb):
    cookie = request.get_cookie('auth')
    
    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie},{'user':1})['user']
        upload = request.files.get('datafile')

        if upload is None:
            redirect('/datos/?data_error=-1')

        name, ext = os.path.splitext(upload.filename)
        datatype = request.forms.get('datatype')

        if ext not in ('.xlsx') or datatype is None:
            redirect('/datos/?data_error=1')

        try:
            wb = load_workbook(filename=upload.file)
        except Exception:
            redirect('/datos/?data_error=2')

        ws = wb.active
        
        if datatype in ('cecps', 'geopos', 'params', 'layers'):
            if datatype == 'cecps':
                # Check for empty or incomplete data
                if ws['A2'].value is None:
                    redirect('/datos/?data_error=41')
                elif ws['B2'].value is None:
                    redirect('/datos/?data_error=42')
                elif ws['C2'].value is None:
                    redirect('/datos/?data_error=43')

            elif datatype == 'geopos':
                # Check for extra data in wrong format
                if ws['E2'].value is not None:
                    redirect('/datos/?data_error=51')
                elif ws['F2'].value is not None:
                    redirect('/datos/?data_error=52')
                elif ws['G2'].value is not None:
                    redirect('/datos/?data_error=53')

            elif datatype == 'params':
                # Check for extra data in wrong format
                if ws['C2'].value is not None:
                    redirect('/datos/?data_error=61')
                elif ws['D2'].value is not None:
                    redirect('/datos/?data_error=62')
                elif ws['E2'].value is not None:
                    redirect('/datos/?data_error=63')

            elif datatype == 'layers':
                # Must check for wrong format
                if ws['A2'].value is None:
                    redirect('/datos/?data_error=71')
                elif ws['B2'].value is None:
                    redirect('/datos/?data_error=72')
                elif ws['C2'].value is None:
                    redirect('/datos/?data_error=73')
                elif ws['D2'].value is None:
                    redirect('/datos/?data_error=74')
                elif ws['E2'].value is not None:
                    redirect('/datos/?data_error=75')

            saveData(mongodb, user, ws, datatype)
        else:
            redirect('/datos/?data_error=3')

        redirect('/datos/?data_error=0')
    else:
        redirect('/')

@app.post('/canal/crear')
def make_chanel(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie},{'user':1})['user']
        channelYear = int(request.forms.get('channelYear'))
        channelYears = int(request.forms.get('channelYears'))
        channelMethod = request.forms.get('channelMethod')

        yearsData = getDataByWeeks(mongodb, user)
        filterYears = np.zeros([channelYears,52])
        populations = {'base':0,'others':np.zeros([channelYears])}
        baseYear = []

        i = 0
        for n in yearsData:
            year = int(yearsData[n]['year'])

            if year == channelYear:
                baseYear = loads(yearsData[n]['data'])
                populations['base'] = int(yearsData[n]['population'])
            elif year < channelYear and i < channelYears:
                filterYears[i,] = np.array(loads(yearsData[n]['data']))
                populations['others'][i] = int(yearsData[n]['population'])
                i += 1

        if i < channelYears:
            return {'error':'No hay datos suficientes para construir el canal.'}
        else:
            return getChannelData(baseYear, filterYears, populations, channelMethod, channelYears)
    else:
        redirect('/')

@app.post('/data/remove')
def delete_year_from_db(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie}, {'user':1})['user']
        yearId = str(request.forms.get('yearId'))
        result = mongodb[user].byWeeks.data.delete_one({'_id':ObjectId(yearId)})
        return str(result.deleted_count)
    else:
        redirect('/')

@app.get('/data/download/<yearId>')
def download_year_from_db(mongodb, yearId):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie):
        user = mongodb.sessions.find_one({'cookie':cookie}, {'user':1})['user']
        try:
            result = mongodb[user].byWeeks.data.find_one({'_id':ObjectId(yearId)})
            if result is None:
                redirect('/datos/?error=404')
        except Exception:
            redirect('/datos/?error=500')

        year = result['year']
        population = result['population']
        data = loads(result['data'])

        wb = load_workbook(filename=data_download_template)
        ws = wb.active
        ws['A2'].value = year
        ws['B2'].value = population

        i = 0
        for row in ws.iter_cols(min_row=2, min_col=3, max_row=2, max_col=54):
            for cell in row:
                cell.value = data[i]
                i += 1

        filename = 'datos_'+year+'.xlsx'
        file_dir = 'static/files/'

        wb.save(file_dir+filename)
        return static_file(filename, root=file_dir, download=filename)
    else:
        redirect('/')

@app.post('/simulador/sim')
def do_simulation(mongodb):
    params = request.forms.get('params')
    return dengue(loads(params))

@app.post('/login')
def check_pass(mongodb):
    username = request.forms.get('username')
    password = request.forms.get('password')

    userRole = check_login(mongodb, username, password)

    if userRole > 0:
        cookie = 'r'+str(userRole)+str(random()).split('.')[1]
        mongodb.sessions.insert_one({'user':username, 'cookie':cookie, 'created_at':datetime.utcnow()})
        response.set_cookie('auth', cookie, max_age=31536000)
        redirect('/')
    else:
        redirect('/ingreso/?login_error=1')

@app.route('/usuarios/')
def users(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie) and cookie[1] == "3":
        return tpl('admin/users', 'Admin | Usuarios', data=getUsers(mongodb))
    else:
        redirect('/')

@app.route('/usuarios/nuevo')
def users(mongodb):
    cookie = request.get_cookie('auth')

    if is_auth(mongodb, cookie) and cookie[1] == "3":
        try:
            error_flag = request.query['new_error']
            if error_flag == '1':
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>Por favor revisa de nuevo los datos que ingresaste.</li></ul></div>'
            elif error_flag == '2':
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>Las contraseñas no coinciden.</li></ul></div>'
            elif error_flag == '3':
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>El usuario que ingresaste ya existe.</li></ul></div>'
            elif error_flag == '0':
                error_msg = '<div style="display:block;" class="ui success message"><ul class="list"><li>El usuario fue creado correctamente.</li></ul></div>'
            else:
                error_msg = ''

        except KeyError:
            error_flag = False
            error_msg = ''

        return tpl('admin/new', 'Admin | Nuevo', error_msg)
    else:
        redirect('/')

@app.post('/usuarios/crear')
def new_user(mongodb):
    user = request.forms.get('user')
    pass1 = request.forms.get('pass1')
    pass2 = request.forms.get('pass2')
    role = request.forms.get('role')

    cookie = request.get_cookie('auth')

    if cookie is None or cookie[1] != "3":
        redirect('/')

    elif (user == '' or pass1 == '' or pass2 == '' or role == ''):
        redirect('/usuarios/nuevo?new_error=1')

    elif (pass1 != pass2):
        redirect('/usuarios/nuevo?new_error=2')

    elif(mongodb.users.count({'user':user}) == 0):
        pass_hash = bcrypt.hashpw(pass1.encode('utf8'), bcrypt.gensalt())
        mongodb.users.insert_one({'user':user,'hash':pass_hash, 'role':role})
        redirect('/usuarios/nuevo?new_error=0')
    else:
        redirect('/usuarios/nuevo?new_error=3')

@app.post('/usuarios/actualizar')
def users_update_post(mongodb):
    user = request.forms.get('user')
    role = request.forms.get('role')
    cookie = request.get_cookie('auth')

    if user is None or role is None:
        redirect('/usuarios/')

    if is_auth(mongodb, cookie) and cookie[1] == "3":
        return tpl('admin/update', 'Admin | Actualizar', '', data={'user':user,'role':role})
    else:
        redirect('/')

@app.get('/usuarios/actualizar')
def users_update_get(mongodb):
    cookie = request.get_cookie('auth')

    try:
        user = request.query['user']
        role = request.query['role']
    except KeyError:
        redirect('/usuarios/')

    if is_auth(mongodb, cookie) and cookie[1] == "3":
        try:
            error_flag = request.query['update_error']
            if error_flag == '1':
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>Por favor revisa de nuevo los datos que ingresaste.</li></ul></div>'
            elif error_flag == '2':
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>Las contraseñas no coinciden.</li></ul></div>'
            elif error_flag == '3':
                error_msg = '<div style="display:block;" class="ui error message"><ul class="list"><li>El usuario que ingresaste no existe.</li></ul></div>'
            elif error_flag == '0':
                error_msg = '<div style="display:block;" class="ui success message"><ul class="list"><li>El usuario fue actualizado correctamente.</li></ul></div>'
            else:
                error_msg = ''

        except KeyError:
            error_flag = False
            error_msg = ''

        return tpl('admin/update', 'Admin | Actualizar', error_msg, data={'user':user,'role':role})
    else:
        redirect('/')

@app.post('/usuarios/modificar')
def change_user(mongodb):
    user = request.forms.get('_user')
    pass1 = request.forms.get('pass1')
    pass2 = request.forms.get('pass2')
    role = request.forms.get('role')

    cookie = request.get_cookie('auth')

    if cookie is None or cookie[1] != "3":
        redirect('/')

    elif (user == '' or pass1 == '' or pass2 == '' or role == ''):
        redirect('/usuarios/actualizar?update_error=1&user=%s&role=%s'% (user,role))

    elif (pass1 != pass2):
        redirect('/usuarios/actualizar?update_error=2&user=%s&role=%s'% (user,role))

    elif (mongodb.users.count({'user':user}) == 1):
        pass_hash = bcrypt.hashpw(pass1.encode('utf8'), bcrypt.gensalt())
        mongodb.users.update_one({'user':user},{'$set':{'hash':pass_hash, 'role':role}})
        redirect('/usuarios/actualizar?update_error=0&user=%s&role=%s'% (user,role))
    else:
        redirect('/usuarios/actualizar?update_error=3&user=%s&role=%s'% (user,role))

@app.post('/usuarios/cambiar')
def username_change(mongodb):
    cookie = request.get_cookie('auth')

    if(is_auth(mongodb, cookie)):
        newUser = request.forms.get('newUser')
        try:
            role = mongodb.users.find_one({'user':newUser},{'user':1,'role':1})['role']
        except Exception:
            role = None
        if(role == 'editor'):
            mongodb.sessions.update_one({'cookie':cookie},{'$set':{'user':newUser}})
            return '1'

    return ''

@app.post('/usuarios/eliminar')
def remove_user(mongodb):
    id = request.forms.get('id')
    cookie = request.get_cookie('auth')

    if cookie is None or cookie[1] != "3":
        redirect('/')

    elif(id == ''):
        redirect('/usuarios/?remove_error=1')

    elif(mongodb.users.count({'_id':ObjectId(id)}) == 0):
        redirect('/usuarios/?remove_error=2')

    else:
        mongodb.users.remove({'_id':ObjectId(id)})
        redirect('/usuarios/')

@app.route('/salir')
def remove_auth(mongodb):
    cookie = request.get_cookie('auth')

    if cookie is not None:
        mongodb.sessions.delete_one({'cookie':cookie})
        response.set_cookie('auth', '', max_age=0)

    redirect('/')

@app.route('/imgs/<filename:path>')
def send_css(filename):
    response = static_file(filename, root='static/imgs')
    response.set_header("Cache-Control", "public, max-age=604800")
    #response.set_header("Cache-Control", "public, max-age=0")
    return response

@app.route('/css/<filename:path>')
def send_css(filename):
    response = static_file(filename, root='static/css')
    #response.set_header("Cache-Control", "public, max-age=604800")
    response.set_header("Cache-Control", "public, max-age=0")
    return response

@app.route('/js/<filename>')
def send_js(filename):
    response = static_file(filename, root='static/js')
    response.set_header("Cache-Control", "public, max-age=0")
    return response

@app.route('/plantillas/<filename:path>')
def download(filename):
    return static_file(filename, root='static/files/templates', download=filename)

if debugging and os.environ.get('APP_LOCATION') == 'LOCAL':
    run(app,host='0.0.0.0', server='wsgiref', plugins=[plugin], port=5000, debug=True, reloader=True)
else:
    run(app,host='0.0.0.0', server='gunicorn', workers=4, plugins=[plugin], port=os.environ.get('PORT'), quiet=True)
